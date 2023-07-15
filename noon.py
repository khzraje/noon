import telebot
from openpyxl import Workbook, load_workbook
import logging

# Set up logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# تكوين البوت باستخدام توكن البوت الخاص بك
bot = telebot.TeleBot("5899737775:AAEXdzYnSI4GFpBVBehK8s7ehhdiypcvtC8")

# قاموس لتخزين معلومات المستخدمين
users = {}


@bot.message_handler(commands=['start'])
def send_welcome(message):
    user_id = message.from_user.id
    if check_user_existence(user_id):
        bot.reply_to(message, "مرحبًا بك مرة أخرى! لقد قمت بالتسجيل بالفعل ✅.")
        send_group_link_active(message)
        group_link = "https://t.me/+vFdcgQFt0UJlNWFi"
        bot.send_message(message.chat.id, f"هذا هو رابط المجموعة : {group_link}")

    else:
        bot.reply_to(message, "مرحبًا بك! يرجى إرسال كلمة 'تسجيل' لبدء التسجيل واستعلام المعلومات.")
        users[message.from_user.id] = {}
        bot.register_next_step_handler(message, process_registration)


def check_user_existence(user_id):
    with open('user_ids.txt', 'r') as file:
        user_ids = file.readlines()
        user_ids = [int(uid.strip()) for uid in user_ids]

    return user_id in user_ids

def send_group_link_active(message):
    group_link = "https://t.me/+vFdcgQFt0UJlNWFi"
def send_group_link(message):
    group_link = "https://t.me/+vFdcgQFt0UJlNWFi"
    bot.send_message(message.chat.id, f"هذا رابط المجموعة، قم بطلب الانضمام للمجموعة وارسل مبلغ قدره دولار واحد 1 الى الرقم 07715648424 لتفعيل عضويتك: {group_link}")


def process_registration(message):
    if message.text == "تسجيل":
        bot.reply_to(message, "شكرًا لاهتمامك بالتسجيل! يرجى إرسال الاسم الثلاثي.")
        users[message.from_user.id] = {}
        bot.register_next_step_handler(message, get_full_name)
    else:
        bot.reply_to(message, "مرحبًا بك! يرجى إرسال كلمة 'تسجيل' لبدء التسجيل واستعلام المعلومات.")
        users[message.from_user.id] = {}
        bot.register_next_step_handler(message, process_registration)
        

@bot.message_handler(func=lambda message: True)
def handle_message(message):
    user_id = message.from_user.id
    if not check_user_existence(user_id):
        bot.reply_to(message, "يرجى إرسال كلمة 'تسجيل' لبدء التسجيل واستعلام المعلومات.")
        bot.register_next_step_handler(message, process_registration)
    else:
        bot.reply_to(message, "مرحبًا بك مرة أخرى! لقد قمت بالتسجيل بالفعل ✅.")
        send_group_link_active(message)
        group_link = "https://t.me/+vFdcgQFt0UJlNWFi"
        bot.send_message(message.chat.id, f"هذا هو رابط المجموعة : {group_link}")

def get_full_name(message):
    try:
        user = users[message.from_user.id]
        user['full_name'] = message.text
        bot.reply_to(message, "من فضلك قم بإرسال العمر.")
        bot.register_next_step_handler(message, get_age)
    except Exception as e:
        bot.reply_to(message, "حدث خطأ في جمع المعلومات.")


def get_age(message):
    try:
        user = users[message.from_user.id]
        age = message.text

        if not age.isdigit() or int(age) < 1 or int(age) > 100:
            error_message = "يجب أن يكون العمر رقمًا بين 1 و 100. يرجى إعادة إرسال العمر."
            bot.reply_to(message, error_message)
            bot.register_next_step_handler(message, get_age)  # تسجيل الخطوة التالية للتحقق من العمر مرة أخرى
            return

        user['age'] = age
        bot.reply_to(message, "من فضلك قم بإرسال عنوان السكن.")
        bot.register_next_step_handler(message, get_address)

    except Exception as e:
        bot.reply_to(message, "حدث خطأ في جمع المعلومات.")



def get_address(message):
    try:
        user = users[message.from_user.id]
        user['address'] = message.text
        bot.reply_to(message, "من فضلك قم بإرسال رقم الهاتف.")
        bot.register_next_step_handler(message, get_phone_number)
    except Exception as e:
        bot.reply_to(message, "حدث خطأ في جمع المعلومات.")


def get_phone_number(message):
    try:
        user = users[message.from_user.id]
        phone_number = message.text

        if not phone_number.isdigit() or len(phone_number) != 11:
            error_message = "يجب أن يتكون رقم الهاتف من 11 رقمًا وأن يحتوي على أرقام فقط. يرجى إعادة إرسال رقم الهاتف."
            bot.reply_to(message, error_message)
            bot.register_next_step_handler(message, get_phone_number)  # تسجيل الخطوة التالية للتحقق من رقم الهاتف مرة أخرى
            return

        user['phone_number'] = phone_number
        bot.reply_to(message, "من فضلك قم بإرسال الجنس 'ذكر 🤵🏻‍♂️' أو 'أنثى 🧕🏻'.")
        bot.register_next_step_handler(message, get_gender)

    except Exception as e:
        bot.reply_to(message, "حدث خطأ في جمع المعلومات.")


# تعريف اسم الملف
excel_file = "data.xlsx"

# دالة لتسجيل معلومات المستخدم في ملف Excel
def save_user_data(user):
    # تحقق مما إذا كان الملف موجودًا أم لا
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Full Name", "Age", "Address", "Phone Number", "Gender"])  # إضافة رأس الجدول

    # قراءة المعلومات من القاموس
    full_name = user["full_name"]
    age = user["age"]
    address = user["address"]
    phone_number = user["phone_number"]
    gender = user["gender"]

    # إضافة المعلومات إلى صف جديد في الجدول
    sheet.append([full_name, age, address, phone_number, gender])

    # حفظ التغييرات في الملف
    workbook.save(excel_file)

# دالة التسجيل النهائية
def register_user(message):
    user_id = message.from_user.id
    if check_user_existence(user_id):
        bot.reply_to(message, "مرحبًا بك مرة أخرى! لقد قمت بالتسجيل بالفعل ✅.")
        send_group_link_active(message)
        group_link = "https://t.me/+vFdcgQFt0UJlNWFi"
        bot.send_message(message.chat.id, f"هذا هو رابط المجموعة : {group_link}")
    else:
        bot.reply_to(message, "مرحبًا بك! يرجى إرسال كلمة 'تسجيل' لبدء التسجيل واستعلام المعلومات.")
        users[message.from_user.id] = {}
        bot.register_next_step_handler(message, process_registration)

# دالة التسجيل النهائية
def get_gender(message):
    try:
        user = users[message.from_user.id]
        gender = message.text

        if gender.lower() not in ['ذكر', 'أنثى']:
            error_message = "الرجاء إرسال 'ذكر 🤵🏻‍♂️' أو 'أنثى 🧕🏻' فقط."
            bot.reply_to(message, error_message)
            bot.register_next_step_handler(message, get_gender)  # تسجيل الخطوة التالية للتحقق من الجنس مرة أخرى
            return

        user['gender'] = gender

        # استدعاء الدالة لحفظ معلومات المستخدم في ملف Excel
        save_user_data(user)

        group_link = "https://t.me/+vFdcgQFt0UJlNWFi"
        user_link = f"https://t.me/{message.from_user.username}"
        bot.send_message(message.chat.id, f"شكرًا لتقديم المعلومات. هذا رابط المجموعة، قم بطلب الانضمام للمجموعة وارسل مبلغ قدره دولار واحد 1 الى الرقم 07715648424 لتفعيل عضويتك: {group_link}")

        # Add user account link to the message
        admin_message = "طلب انضمام جديد:\n\n"
        admin_message += f"الاسم الثلاثي: {user['full_name']}\n"
        admin_message += f"العمر: {user['age']}\n"
        admin_message += f"العنوان: {user['address']}\n"
        admin_message += f"رقم الهاتف: {user['phone_number']}\n"
        admin_message += f"الجنس: {user['gender']}\n"
        admin_message += f"رابط حساب المستخدم: {user_link}\n"

        bot.send_message("929212999", admin_message)  # استبدال ADMIN_CHAT_ID بمعرف الدردشة للمسؤول

        # Save user ID to file
        with open('user_ids.txt', 'a') as file:
            file.write(str(message.from_user.id) + '\n')

    except Exception as e:
        bot.reply_to(message, "حدث خطأ في جمع المعلومات.")


# تشغيل البوت
bot.polling()
